"""Build generated_review.xlsx — the architect's approval sheet for NEW questions.

Reads reports/generated_accepted.json (candidates that passed the auto gate) and
writes one Excel sheet: each row is a full candidate + a verdict dropdown
(Approve / Reject / Needs edit) + notes. `merge_accepted --reviewed` then adds
only the Approve rows to the bank.

Usage (from tools/question_audit/):
    python -m src.review_generated
"""
from __future__ import annotations

import json

from src import config

VERDICTS = ["Approve", "Reject", "Needs edit"]

_ORDER = [
    "REVIEW: verdict", "REVIEW: notes", "id", "section",
    "question", "options (>> = correct)", "correct", "explanation",
    "codeReference", "grounded_on", "repaired", "dup_sim",
]
_WIDTHS = {
    "REVIEW: verdict": 16, "REVIEW: notes": 34, "id": 10, "section": 20,
    "question": 60, "options (>> = correct)": 52, "correct": 38, "explanation": 58,
    "codeReference": 34, "grounded_on": 30, "repaired": 9, "dup_sim": 8,
}
_WRAP = {"REVIEW: notes", "question", "options (>> = correct)", "correct",
         "explanation", "codeReference", "grounded_on"}


def main() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    acc = json.loads((config.REPORTS_DIR / "generated_accepted.json").read_text(encoding="utf-8"))
    if not acc:
        print("No accepted candidates to review.")
        return

    rows = []
    for r in acc:
        opts = "\n".join(
            f"{'>>' if o == r['correctOption'] else '  '} {o}" for o in r["options"]
        )
        rows.append({
            "REVIEW: verdict": "", "REVIEW: notes": "",
            "id": r["id"], "section": r["section"], "question": r["question"],
            "options (>> = correct)": opts, "correct": r["correctOption"],
            "explanation": r["explanation"], "codeReference": r["codeReference"],
            "grounded_on": r.get("grounded_on", ""), "repaired": r.get("repaired", False),
            "dup_sim": r.get("dup_sim", ""),
        })

    dark = PatternFill("solid", fgColor="1F2937")
    yellow = PatternFill("solid", fgColor="F2B807")
    wrap_top = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    ws.append(_ORDER)
    for r in rows:
        ws.append([r[h] for h in _ORDER])

    for c in ws[1]:
        c.fill = dark
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i in (1, 2):
        ws.cell(row=1, column=i).fill = yellow
        ws.cell(row=1, column=i).font = Font(bold=True, color="1F2937")

    last, ncol = len(rows) + 1, len(_ORDER)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{last}"
    for i, h in enumerate(_ORDER, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = _WIDTHS.get(h, 16)
        if h in _WRAP:
            for cell in ws[col][1:]:
                cell.alignment = wrap_top

    helper = ncol + 3
    hcol = get_column_letter(helper)
    for idx, v in enumerate(VERDICTS, start=2):
        ws.cell(row=idx, column=helper, value=v)
    ws.column_dimensions[hcol].hidden = True
    dv = DataValidation(type="list", formula1=f"${hcol}$2:${hcol}${1 + len(VERDICTS)}", allow_blank=True)
    dv.add(f"A2:A{last}")
    ws.add_data_validation(dv)

    out = config.REPORTS_DIR / "generated_review.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = out.with_name("generated_review_new.xlsx")
        wb.save(out)
        print(f"(main file was open) wrote {out.name}")
    print(f"Wrote {out.name}: {len(rows)} candidates to approve/reject.")


if __name__ == "__main__":
    main()
