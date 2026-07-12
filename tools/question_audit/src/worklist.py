"""Build the architect's review sheet — one self-contained, Excel-friendly file.

Each row = the full question (stem, options with the correct one marked,
explanation, citation) + issue tags + the AI's reasons + any duplicate partner.

The flagged questions are triaged into three queues (see `queue_for`). What we
actually hand the architect is the RED queue: the questions where a licensed
architect's judgment is the only thing that can settle them. YELLOW is an
AI-repair-then-approve flow and GREEN is a spot-check, so neither belongs in her
file — sending all 692 rows would be ~23 hours of work, most of it wasted.

Usage (from tools/question_audit/):
    python -m src.worklist --queue RED     # what we send the architect
    python -m src.worklist                 # everything (all three queues)
"""
from __future__ import annotations

import argparse
import csv
import json
from collections import Counter

from src import config
from src.schema import load_questions


# Dropdown choices for the architect's verdict column.
VERDICTS = [
    "OK - оставить",
    "Исправить ответ",
    "Заменить обманки",
    "Переписать",
    "Удалить",
    "Удалить дубликат",
]

# Guide sheet content: (style, text). style in {title, h, body}.
GUIDE = [
    ("title", "Проверка вопросов ARE — инструкция для Марианны"),
    ("body", "Ты практикующий архитектор в Нью-Йорке — твоя экспертиза здесь главное. Писать вопросы с нуля НЕ нужно. Задача: пройти список на вкладке «worklist» и по каждому вопросу выбрать решение в колонке «REVIEW: verdict»."),

    ("h", "Почему здесь только 127 вопросов, а не все 1082"),
    ("body", "Мы прогнали весь банк через AI-проверку. Она отсортировала вопросы на три группы. Тебе мы отдаём ТОЛЬКО ту группу, где AI не имеет права быть судьёй и нужен живой архитектор. Остальное AI чинит сам, а мы потом просто утверждаем."),
    ("body", "Что попало в твой список: (1) вопросы, где ОТВЕТ, ОБЪЯСНЕНИЕ и ССЫЛКА НА НОРМУ противоречат друг другу; (2) самые тяжёлые по совокупности проблем; (3) дубликаты, где надо решить, какой из двух оставить."),
    ("body", "Пример, чтобы было понятно, о чём речь. Вопрос ppd_q3 спрашивает про длиннопролётные конструкции на ГРАВИТАЦИОННУЮ нагрузку, а в ссылке стоит «AISC Seismic Design Manual» — сейсмический мануал. Ответ, возможно, верный, а ссылка — нет. Такое ловится только глазами архитектора."),

    ("h", "Сколько это времени — честно"),
    ("body", "127 вопросов. Это НЕ «по 2 минуты»: почти каждый требует сверить норму или пункт контракта. Реалистично 5-8 минут на вопрос, то есть примерно 11-17 ЧАСОВ. Растяни на несколько дней, иди сверху вниз — самые проблемные наверху."),
    ("body", "Если времени совсем мало: отфильтруй колонку «issues» по CONSISTENCY (их 52) — это самое важное и самое опасное."),

    ("h", "Главный вопрос к каждому пункту"),
    ("body", "«Столкнулся бы реальный архитектор с этим на практике — и правильный ли здесь ответ и ссылка?» Твой практический опыт важнее всего: чаще всего именно ты ловишь неверный «правильный» ответ."),

    ("h", "Что значат метки (колонка issues)"),
    ("body", "CONSISTENCY — САМОЕ ВАЖНОЕ. Ответ, объяснение и ссылка на норму противоречат друг другу. Смотри колонку «consistency_problem» — там написано, что именно нашёл AI. Проверь по коду/практике."),
    ("body", "DUPLICATE — почти такой же вопрос уже есть в банке (см. колонку «duplicate_of»). Реши, какой оставить, второй помечай «Удалить дубликат»."),
    ("body", "FACTUAL — вопрос на знание факта, а не на суждение. ВАЖНО: это НЕ значит «плохой вопрос». Настоящий ARE тоже спрашивает факты. НЕ переписывай вопрос только за то, что он FACTUAL."),
    ("body", "WEAK_DISTRACTORS — неправильные варианты слишком очевидно неверны. LEAKAGE — ответ угадывается по формулировке. Эти две пометки здесь второстепенные: их AI чинит сам, тебе на них отвлекаться не надо."),

    ("h", "Как проверять — по шагам"),
    ("body", "1. Прочитай вопрос и правильный ответ (помечен >> в колонке «options»)."),
    ("body", "2. Ответ действительно верный? Если нет → «Исправить ответ», впиши верный в «REVIEW: fix / notes»."),
    ("body", "3. Ссылка на норму (колонка «codeReference») действительно соответствует вопросу? Если нет — впиши верную ссылку в notes и поставь «Исправить ответ»."),
    ("body", "4. Объяснение не противоречит ответу? Если противоречит — поправь в notes."),
    ("body", "5. Дубликат? → «Удалить дубликат» у того, который убираем."),
    ("body", "6. Всё хорошо → «OK - оставить». Это нормальный и частый исход — AI мог ошибиться, ты его подтверждаешь или опровергаешь."),

    ("h", "Колонка REVIEW: verdict — выбери из выпадающего списка"),
    ("body", "OK - оставить — вопрос хороший, AI поднял ложную тревогу, не трогаем."),
    ("body", "Исправить ответ — вопрос ок, но неверен «правильный» ответ ИЛИ ссылка на норму (верное пиши в notes)."),
    ("body", "Заменить обманки — правильный ответ ок, но неправильные варианты слишком очевидны."),
    ("body", "Переписать — тема нужная, но формулировка плохая."),
    ("body", "Удалить — вопрос плохой или нерелевантный, убрать совсем."),
    ("body", "Удалить дубликат — повтор, убрать."),

    ("h", "Как писать правку"),
    ("body", "В колонку «REVIEW: fix / notes» пиши свободным текстом: верный ответ, верную ссылку на норму или короткий комментарий. Не надо оформлять — мы разберём."),
]

def queue_for(tags: list[str], severity: float) -> str:
    """Triage a flagged question into the queue that decides WHO reviews it.

    RED    - needs a licensed architect's judgment and nothing less: the answer,
             explanation and citation disagree; or the question is among the worst
             scored; or it duplicates another and someone must pick the survivor.
    YELLOW - mechanical weaknesses (throwaway distractors, answer telegraphed by
             the stem). An LLM can propose a fix; the architect only approves it.
    GREEN  - flagged but with no conflict. Spot-check a sample; do not rewrite a
             question merely for being factual -- the real ARE tests facts too.
    """
    if "CONSISTENCY" in tags or "DUPLICATE" in tags or severity >= 7:
        return "RED"
    if "WEAK_DISTRACTORS" in tags or "LEAKAGE" in tags:
        return "YELLOW"
    return "GREEN"


# Column order for the worklist sheet — review columns first (frozen).
_ORDER = [
    "REVIEW: verdict", "REVIEW: fix / notes",
    "priority", "severity", "id", "section", "issues",
    "question", "options (>> = correct)", "correct", "explanation", "codeReference",
    "why_factual", "distractor_notes", "consistency_problem", "leakage_note",
    "difficulty", "duplicate_of",
]
_WIDTHS = {
    "REVIEW: verdict": 20, "REVIEW: fix / notes": 34, "priority": 9, "severity": 8, "id": 11, "section": 20,
    "issues": 20, "question": 55, "options (>> = correct)": 50, "correct": 36, "explanation": 58,
    "codeReference": 26, "why_factual": 36, "distractor_notes": 52, "consistency_problem": 34,
    "leakage_note": 34, "difficulty": 10, "duplicate_of": 16,
}
_WRAP = {
    "REVIEW: fix / notes", "question", "options (>> = correct)", "correct", "explanation",
    "codeReference", "why_factual", "distractor_notes", "consistency_problem", "leakage_note",
}


def _write_xlsx(rows: list[dict], path) -> None:
    """Two-sheet workbook: a Russian guide, then the review sheet with a verdict
    dropdown + notes column, frozen header/review-columns, autofilter, wrap."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    dark = PatternFill("solid", fgColor="1F2937")
    yellow = PatternFill("solid", fgColor="F2B807")
    wrap_top = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    # ── Sheet 1: guide ───────────────────────────────────────────────────────
    g = wb.active
    g.title = "Инструкция"
    g.sheet_properties.tabColor = "F2B807"
    g.column_dimensions["A"].width = 118
    for i, (style, text) in enumerate(GUIDE, 1):
        c = g.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if style == "title":
            c.font = Font(bold=True, size=14, color="1F2937")
        elif style == "h":
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = dark
        else:
            c.font = Font(size=10)

    # ── Sheet 2: worklist ────────────────────────────────────────────────────
    ws = wb.create_sheet("worklist")
    ws.append(_ORDER)
    for r in rows:
        ws.append(["" if h.startswith("REVIEW") else r.get(h, "") for h in _ORDER])

    for c in ws[1]:
        c.fill = dark
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for i in (1, 2):  # highlight the two review columns in the header
        hc = ws.cell(row=1, column=i)
        hc.fill = yellow
        hc.font = Font(bold=True, color="1F2937")

    last = len(rows) + 1
    ncol = len(_ORDER)
    ws.freeze_panes = "C2"  # keep header row + the two review columns visible
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{last}"
    for i, h in enumerate(_ORDER, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = _WIDTHS.get(h, 16)
        if h in _WRAP:
            for cell in ws[col][1:]:
                cell.alignment = wrap_top

    # Verdict dropdown — values live in a hidden helper column on this sheet.
    helper = ncol + 3
    hcol = get_column_letter(helper)
    for idx, v in enumerate(VERDICTS, start=2):
        ws.cell(row=idx, column=helper, value=v)
    ws.column_dimensions[hcol].hidden = True
    dv = DataValidation(
        type="list", formula1=f"${hcol}$2:${hcol}${1 + len(VERDICTS)}", allow_blank=True
    )
    dv.add(f"A2:A{last}")
    ws.add_data_validation(dv)

    try:
        wb.save(path)
    except PermissionError:
        alt = path.with_name(path.stem + "_new.xlsx")
        wb.save(alt)
        print(f"NOTE: {path.name} is open in Excel (locked) — wrote {alt.name} instead. "
              f"Close Excel and re-run to update the main file.")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--min-severity", type=float, default=4.0)
    ap.add_argument(
        "--queue",
        choices=["RED", "YELLOW", "GREEN", "ALL"],
        default="ALL",
        help="RED = the only queue we actually hand the architect",
    )
    args = ap.parse_args()

    audit = {
        r["id"]: r
        for r in json.loads((config.REPORTS_DIR / "audit_1100.json").read_text(encoding="utf-8"))[
            "results"
        ]
    }
    qs = {q.id: q for q in load_questions(config.QUESTIONS_PATH).valid}

    # Duplicate partners from the 0.90 list.
    dup: dict[str, list[str]] = {}
    dpath = config.REPORTS_DIR / "duplicates_090.csv"
    if dpath.exists():
        with dpath.open(encoding="utf-8") as f:
            for row in csv.DictReader(f):
                a, b, s = row["id_a"], row["id_b"], row["similarity"]
                dup.setdefault(a, []).append(f"{b} ({s})")
                dup.setdefault(b, []).append(f"{a} ({s})")

    rows = []
    for qid, r in audit.items():
        q = qs.get(qid)
        if not q:
            continue
        d = r["distractors"]["distractors"]
        davg = sum(x["plausibility"] for x in d) / len(d) if d else 0

        tags = []
        if r["judgment"]["label"] == "factual":
            tags.append("FACTUAL")
        if davg < 2.5:
            tags.append("WEAK_DISTRACTORS")
        if r["leakage"]["leakage_prone"]:
            tags.append("LEAKAGE")
        if r["consistency"]["verdict"] == "fail":
            tags.append("CONSISTENCY")
        if qid in dup:
            tags.append("DUPLICATE")

        # Keep if it clears the bar OR has a hard flag we always want reviewed.
        hard = "DUPLICATE" in tags or "CONSISTENCY" in tags
        if r["severity"] < args.min_severity and not hard:
            continue

        opts = "\n".join(f"{'>>' if o == q.correctOption else '  '} {o}" for o in q.options)
        dnotes = " | ".join(
            f"{x['plausibility']}/5 [{x['option'][:22]}]: {x['reason'][:55]}" for x in d
        )
        lk = r["leakage"]
        lnote = (
            f"answerable from stem alone: '{lk['model_answer'][:45]}' (conf {lk['confidence']}/5)"
            if lk["leakage_prone"]
            else ""
        )
        rows.append(
            {
                "priority": queue_for(tags, r["severity"]),
                "severity": r["severity"],
                "id": qid,
                "section": q.section,
                "difficulty": q.difficulty,
                "issues": ",".join(tags),
                "duplicate_of": "; ".join(dup.get(qid, [])),
                "question": q.question,
                "options (>> = correct)": opts,
                "correct": q.correctOption,
                "explanation": q.explanation,
                "codeReference": q.codeReference,
                "why_factual": r["judgment"]["reason"] if "FACTUAL" in tags else "",
                "distractor_notes": dnotes,
                "consistency_problem": r["consistency"]["reason"] if "CONSISTENCY" in tags else "",
                "leakage_note": lnote,
            }
        )

    # RED first, then worst-scored within each queue.
    _RANK = {"RED": 0, "YELLOW": 1, "GREEN": 2}
    rows.sort(key=lambda x: (_RANK[x["priority"]], -x["severity"]))

    counts = Counter(x["priority"] for x in rows)
    red, yellow, green = counts["RED"], counts["YELLOW"], counts["GREEN"]
    print(
        f"Queues: RED {red} (architect only) | YELLOW {yellow} (AI proposes, architect approves) | "
        f"GREEN {green} (spot-check ~10-15%)"
    )

    if args.queue != "ALL":
        rows = [r for r in rows if r["priority"] == args.queue]
        if not rows:
            print(f"No rows in the {args.queue} queue.")
            return

    # RED items each need a code section or a contract clause checked against the
    # source -- that is 5-8 minutes of an architect's time, not 2.
    if args.queue == "RED":
        stem, lo, hi = "maryana_red_queue", round(len(rows) * 5 / 60), round(len(rows) * 8 / 60)
        print(f"RED queue for the architect: {len(rows)} questions, realistically ~{lo}-{hi} h.")
    else:
        stem = "maryana_worklist"
        print(f"All {len(rows)} rows at ~2 min each would be ~{round(len(rows) * 2 / 60)} h -- don't send that.")

    out = config.REPORTS_DIR / f"{stem}.csv"
    with out.open("w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig = clean in Excel
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)

    _write_xlsx(rows, config.REPORTS_DIR / f"{stem}.xlsx")

    tally: Counter = Counter()
    for r in rows:
        for t in r["issues"].split(","):
            if t:
                tally[t] += 1
    print(f"Wrote {stem}.csv / {stem}.xlsx: {len(rows)} questions.")
    print("Issue tag tally:", dict(tally))


if __name__ == "__main__":
    main()
