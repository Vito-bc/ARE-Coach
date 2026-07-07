"""Phase 2 — merge accepted generated questions into the bank.

Reads reports/generated_accepted.json (candidates that passed the auto gate),
assigns fresh 'gen_qN' ids, keeps only the schema fields, validates each, and
appends to questions_ny.json. Backs up first; dry-run by default.

INTENDED FLOW: generate -> gate -> a HUMAN (architect) reviews the accepted
sample -> merge. Only run --apply on questions a human has approved.

Usage (from tools/question_audit/):
    python -m src.merge_accepted            # dry-run
    python -m src.merge_accepted --apply
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime

from src import config
from src.schema import Question, load_questions

_FIELDS = [
    "id", "state", "section", "difficulty", "question", "options",
    "correctOption", "explanation", "codeReference", "examWeight", "topic",
]


def _gen_ids(existing: set[str]):
    """Yield fresh, non-colliding gen_qN ids continuing past any existing ones."""
    n = max((int(m.group(1)) for i in existing if (m := re.fullmatch(r"gen_q(\d+)", i))), default=0)
    while True:
        n += 1
        yield f"gen_q{n}"


def _approved_ids() -> set[str]:
    """Read generated_review.xlsx and return the ids marked 'Approve'."""
    from openpyxl import load_workbook

    path = config.REPORTS_DIR / "generated_review.xlsx"
    if not path.exists():
        print(f"No {path.name} — run `python -m src.review_generated` and have the architect fill it.")
        return set()
    ws = load_workbook(path, read_only=True).active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    vcol, icol = header.index("REVIEW: verdict"), header.index("id")
    approved = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[vcol] and str(row[vcol]).strip().lower() == "approve":
            approved.add(row[icol])
    return approved


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--input", default="generated_accepted.json")
    ap.add_argument("--reviewed", action="store_true",
                    help="merge only rows marked 'Approve' in generated_review.xlsx")
    args = ap.parse_args()

    src_path = config.REPORTS_DIR / args.input
    accepted = json.loads(src_path.read_text(encoding="utf-8")) if src_path.exists() else []
    if not accepted:
        print(f"No accepted candidates in {src_path.name}.")
        return

    if args.reviewed:
        approved = _approved_ids()
        before = len(accepted)
        accepted = [r for r in accepted if r.get("id") in approved]
        print(f"Human review: {len(approved)} marked Approve -> {len(accepted)}/{before} candidates pass.")
        if not accepted:
            print("Nothing approved to merge.")
            return

    bank = json.loads(config.QUESTIONS_PATH.read_text(encoding="utf-8-sig"))
    ids = _gen_ids({q["id"] for q in bank})

    new_rows = []
    for rec in accepted:
        row = {k: rec.get(k) for k in _FIELDS}
        row["id"], row["state"] = next(ids), "NY"
        try:
            Question(**row)  # validate against the schema
        except Exception as e:
            print(f"  SKIP invalid candidate ({e})")
            continue
        new_rows.append(row)

    print(f"Bank {len(bank)} + {len(new_rows)} accepted -> {len(bank) + len(new_rows)}")
    for r in new_rows:
        print(f"  + {r['id']} [{r['section']}] {r['question'][:60]}")
    print("\nNOTE: these passed the AUTO gate only — a human (architect) should review before shipping.")

    if args.apply:
        backups = config.TOOL_DIR / "backups"
        backups.mkdir(exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = backups / f"questions_ny_{ts}_premerge.json"
        shutil.copy2(config.QUESTIONS_PATH, bak)
        config.QUESTIONS_PATH.write_text(
            json.dumps(bank + new_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        res = load_questions(config.QUESTIONS_PATH)
        print(f"BACKED UP -> backups/{bak.name}; wrote {len(bank) + len(new_rows)}. "
              f"Valid: {len(res.valid)}, invalid: {len(res.errors)}")
    else:
        print("DRY-RUN — nothing written. Re-run with --apply.")


if __name__ == "__main__":
    main()
